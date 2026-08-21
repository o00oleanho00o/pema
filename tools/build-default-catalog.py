import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "danhsach" / "danhsach.xlsx"
OUTPUT = ROOT / "default-catalog.json"


def clean(value):
    return "" if value is None else str(value).strip()


workbook = load_workbook(SOURCE, read_only=True, data_only=True)
sheet = workbook[workbook.sheetnames[0]]
rows = []
for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    code, name, unit, source_type = map(clean, (values[1], values[2], values[3], values[4]))
    if not any((code, name, unit, source_type)):
        continue
    rows.append({"code": code, "name": name, "unit": unit, "sourceType": source_type, "rowNumber": row_number})

OUTPUT.write_text(json.dumps({"filename": SOURCE.name, "sheetName": sheet.title, "rows": rows}, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
print(f"Wrote {len(rows)} rows to {OUTPUT}")
